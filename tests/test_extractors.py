"""Tests for extractors.py — extraction and frontmatter parsing for all supported formats."""

from pathlib import Path



# --- Frontmatter parsing ---


def test_parse_frontmatter_basic():
    """Standard YAML frontmatter is parsed and stripped from content."""
    from extractors import parse_frontmatter

    content = "---\ntags: [recipe, korean]\nstatus: active\ncreated: 2026-01-15\n---\n# Hello\n\nBody text."
    fm, body = parse_frontmatter(content)
    assert fm["tags"] == ["recipe", "korean"]
    assert fm["status"] == "active"
    assert "---" not in body
    assert "# Hello" in body
    assert "Body text." in body


def test_parse_frontmatter_no_frontmatter():
    """Content without frontmatter returns empty dict and original content."""
    from extractors import parse_frontmatter

    content = "# Just a heading\n\nNo frontmatter here."
    fm, body = parse_frontmatter(content)
    assert fm == {}
    assert body == content


def test_parse_frontmatter_empty_frontmatter():
    """Empty frontmatter block returns empty dict."""
    from extractors import parse_frontmatter

    content = "---\n---\n# Content"
    fm, body = parse_frontmatter(content)
    # yaml.safe_load("") returns None, which isn't a dict
    assert fm == {}


def test_parse_frontmatter_string_tags():
    """Frontmatter with string tags (not a list)."""
    from extractors import parse_frontmatter

    content = "---\ntags: recipe\n---\n# Recipe"
    fm, body = parse_frontmatter(content)
    assert fm["tags"] == "recipe"


def test_parse_frontmatter_invalid_yaml():
    """Invalid YAML in frontmatter is handled gracefully."""
    from extractors import parse_frontmatter

    content = "---\n[invalid yaml: {\n---\n# Content"
    fm, body = parse_frontmatter(content)
    # Should return original content on parse failure
    assert fm == {}


def test_extract_title_from_heading():
    """Title is extracted from first # heading."""
    from extractors import extract_title

    assert extract_title("# My Great Title\n\nContent", "note.md") == "My Great Title"


def test_extract_title_from_filename():
    """Title falls back to filename when no heading."""
    from extractors import extract_title

    assert extract_title("No heading here, just text.", "my_notes.md") == "my_notes"


def test_extract_title_nested_doc_id():
    """Title from filename works with nested paths."""
    from extractors import extract_title

    assert extract_title("No heading", "subfolder/recipe.md") == "recipe"


def test_derive_folder():
    """Folder is derived from top-level directory."""
    from extractors import derive_folder

    assert derive_folder("subfolder/recipe.md") == "subfolder"
    assert derive_folder("Archive/old_note.md") == "Archive"
    assert derive_folder("note1.md") == ""
    assert derive_folder("a/b/c/deep.md") == "a"


def test_normalize_tags_list():
    """List tags are joined into comma-separated string."""
    from extractors import normalize_tags

    assert normalize_tags(["recipe", "korean"]) == "recipe,korean"


def test_normalize_tags_string():
    """String tags are normalized."""
    from extractors import normalize_tags

    assert normalize_tags("recipe, korean, fermentation") == "recipe,korean,fermentation"


def test_normalize_tags_none():
    """None tags return empty string."""
    from extractors import normalize_tags

    assert normalize_tags(None) == ""


def test_normalize_tags_single():
    """Single string tag."""
    from extractors import normalize_tags

    assert normalize_tags("recipe") == "recipe"


# --- Markdown extraction ---


def test_extract_markdown(tmp_path):
    """Markdown extraction returns the file contents."""
    from extractors import extract_markdown

    md_file = tmp_path / "test.md"
    md_file.write_text("# Hello\n\nSome content here.", encoding="utf-8")

    result = extract_markdown(md_file)
    assert "Hello" in result.full_text
    assert "Some content here" in result.full_text
    assert len(result.pages) == 1
    assert result.pages[0].page == 0


def test_extract_markdown_with_frontmatter(tmp_path):
    """Markdown with frontmatter: YAML is parsed, content is stripped of frontmatter."""
    from extractors import extract_markdown

    md_file = tmp_path / "with_fm.md"
    md_file.write_text(
        "---\ntags: [project, AI]\nstatus: active\ncreated: 2026-01-10\n---\n"
        "# My Project\n\nProject description here.",
        encoding="utf-8",
    )

    result = extract_markdown(md_file)
    # Frontmatter should be parsed
    assert result.frontmatter["tags"] == ["project", "AI"]
    assert result.frontmatter["status"] == "active"
    # Content should NOT contain the YAML block
    assert "---" not in result.full_text
    assert "tags:" not in result.full_text
    # Content should have the actual text
    assert "# My Project" in result.full_text
    assert "Project description" in result.full_text


def test_extract_markdown_no_frontmatter(tmp_path):
    """Markdown without frontmatter: frontmatter dict is empty."""
    from extractors import extract_markdown

    md_file = tmp_path / "no_fm.md"
    md_file.write_text("# Plain Note\n\nJust text.", encoding="utf-8")

    result = extract_markdown(md_file)
    assert result.frontmatter == {}
    assert "# Plain Note" in result.full_text


def test_extract_markdown_empty(tmp_path):
    """Empty markdown file returns empty text."""
    from extractors import extract_markdown

    md_file = tmp_path / "empty.md"
    md_file.write_text("", encoding="utf-8")

    result = extract_markdown(md_file)
    assert result.full_text == ""


def test_extract_markdown_unicode(tmp_path):
    """Markdown with Unicode characters."""
    from extractors import extract_markdown

    md_file = tmp_path / "unicode.md"
    md_file.write_text("日本語テスト 🚀 café", encoding="utf-8")

    result = extract_markdown(md_file)
    assert "日本語テスト" in result.full_text
    assert "🚀" in result.full_text


# --- PDF extraction ---


def _make_test_pdf(path: Path, pages_text: list[str]) -> None:
    """Create a test PDF with the given text on each page."""
    import fitz

    doc = fitz.open()
    for text in pages_text:
        page = doc.new_page()
        page.insert_text((72, 100), text, fontsize=12)
    doc.save(str(path))
    doc.close()


def test_extract_pdf_text_only(tmp_path):
    """PDF text_only strategy extracts native text without OCR."""
    from extractors import extract_pdf

    pdf_path = tmp_path / "test.pdf"
    _make_test_pdf(pdf_path, ["Page one content", "Page two content"])

    result = extract_pdf(pdf_path, strategy="text_only")
    assert "Page one" in result.full_text
    assert "Page two" in result.full_text
    assert len(result.pages) == 2
    assert result.pages[0].page == 0
    assert result.pages[1].page == 1
    assert not any(p.was_ocr for p in result.pages)


def test_extract_pdf_encrypted_is_skipped(tmp_path):
    """A password-protected PDF opens but raises on page read; it must be flagged
    as an 'encrypted_pdf' skip (so it lands in the skip ledger and isn't
    re-extracted/re-thrown every run) rather than bubbling an exception."""
    import fitz
    from extractors import begin_degradation_capture, collect_skips, extract_pdf

    pdf_path = tmp_path / "encrypted.pdf"
    doc = fitz.open()
    doc.new_page().insert_text((72, 100), "secret content", fontsize=12)
    doc.save(str(pdf_path), encryption=fitz.PDF_ENCRYPT_AES_256,
             user_pw="pw", owner_pw="pw")
    doc.close()

    begin_degradation_capture()
    result = extract_pdf(pdf_path, strategy="text_then_ocr")

    assert result.full_text == ""
    assert "encrypted_pdf" in collect_skips()


def test_extract_pdf_text_then_ocr_sufficient_text(tmp_path):
    """text_then_ocr doesn't call OCR when there's enough native text."""
    import fitz
    from extractors import extract_pdf

    # Create a PDF with enough text (insert multiple lines to exceed 200 chars)
    pdf_path = tmp_path / "test.pdf"
    doc = fitz.open()
    page = doc.new_page()
    y = 72
    for i in range(20):
        page.insert_text((72, y), f"Line {i}: This is enough text content to fill the page properly.", fontsize=11)
        y += 16
    doc.save(str(pdf_path))
    doc.close()

    result = extract_pdf(pdf_path, strategy="text_then_ocr", ocr_provider=None, min_text_chars=200)
    assert len(result.full_text) > 200
    assert not result.pages[0].was_ocr


def test_extract_pdf_text_then_ocr_calls_ocr_on_thin_page(tmp_path):
    """text_then_ocr falls back to OCR when native text is too short."""
    from extractors import extract_pdf
    from providers.ocr.base import OCRProvider

    class FakeOCR(OCRProvider):
        def __init__(self):
            self.called = False
        def extract(self, file_path, page=None):
            self.called = True
            return "OCR extracted text from image"

    pdf_path = tmp_path / "test.pdf"
    _make_test_pdf(pdf_path, ["Hi"])  # Very short text

    fake_ocr = FakeOCR()
    result = extract_pdf(pdf_path, strategy="text_then_ocr", ocr_provider=fake_ocr, min_text_chars=200)
    assert fake_ocr.called
    assert result.pages[0].was_ocr
    assert "OCR extracted text" in result.full_text


def test_extract_pdf_respects_ocr_page_limit(tmp_path):
    """OCR is not called beyond ocr_page_limit."""
    from extractors import extract_pdf
    from providers.ocr.base import OCRProvider

    class CountingOCR(OCRProvider):
        def __init__(self):
            self.call_count = 0
        def extract(self, file_path, page=None):
            self.call_count += 1
            return "OCR text"

    pdf_path = tmp_path / "test.pdf"
    _make_test_pdf(pdf_path, ["Hi"] * 5)  # 5 pages, all short

    counting_ocr = CountingOCR()
    result = extract_pdf(pdf_path, strategy="text_then_ocr", ocr_provider=counting_ocr,
                         min_text_chars=200, ocr_page_limit=2)
    # Only first 2 pages should be OCR'd
    assert counting_ocr.call_count == 2


def test_extract_pdf_single_empty_page(tmp_path):
    """PDF with one blank page returns only metadata header (no document text)."""
    import fitz
    from extractors import extract_pdf

    pdf_path = tmp_path / "blank.pdf"
    doc = fitz.open()
    doc.new_page()  # blank page, no text inserted
    doc.save(str(pdf_path))
    doc.close()

    result = extract_pdf(pdf_path)
    assert len(result.pages) == 1
    assert result.frontmatter.get("page_count") == 1
    # Only metadata header, no actual document text beyond it
    lines = [l for l in result.full_text.strip().splitlines() if not l.startswith(("Pages:", "Author:", "Created:", "Document title:"))]
    assert all(l.strip() == "" for l in lines)


# --- Image extraction ---


def test_extract_image_with_ocr():
    """Image extraction calls OCR provider."""
    from extractors import extract_image
    from providers.ocr.base import OCRProvider

    class FakeOCR(OCRProvider):
        def extract(self, file_path, page=None):
            return "Meeting notes from the image"

    result = extract_image("/fake/path.png", ocr_provider=FakeOCR())
    assert "Meeting notes" in result.full_text


def test_extract_image_no_ocr():
    """Image extraction without OCR returns empty."""
    from extractors import extract_image

    result = extract_image("/fake/path.png", ocr_provider=None)
    assert result.full_text == ""


def test_extract_image_failed_describe_notes_transient_reason():
    from extractors import begin_degradation_capture, collect_degradations, extract_image
    from providers.ocr.base import OCRProvider

    class BoomOCR(OCRProvider):
        def extract(self, file_path, page=None): return ""
        def describe(self, file_path): raise ConnectionError("vision host down")

    begin_degradation_capture()
    extract_image("/fake/path.png", ocr_provider=BoomOCR())
    degs = collect_degradations()
    assert [d.reason for d in degs] == ["ocr_describe_failed"]
    assert degs[0].transient is True  # ConnectionError is transient


def test_extract_image_confirmed_blank_returns_no_degradation():
    # The describe provider is always wrapped: it returns "" ONLY for a
    # fallback-confirmed blank. extract_image must treat that as clean (no
    # degradation note) so the doc drops from the ledger and is never re-described.
    from extractors import begin_degradation_capture, collect_degradations, extract_image
    from providers.ocr.base import OCRProvider

    class BlankOCR(OCRProvider):
        def extract(self, file_path, page=None): return ""
        def describe(self, file_path): return ""

    begin_degradation_capture()
    extract_image("/fake/path.png", ocr_provider=BlankOCR())
    assert collect_degradations() == []


def test_extract_image_good_describe_notes_nothing():
    from extractors import begin_degradation_capture, collect_degradations, extract_image
    from providers.ocr.base import OCRProvider

    class GoodOCR(OCRProvider):
        def extract(self, file_path, page=None):
            return "A receipt for $50"

    begin_degradation_capture()
    extract_image("/fake/path.png", ocr_provider=GoodOCR())
    assert collect_degradations() == []


# --- Audio/video extraction degradations ---
# Same rule as images (#0264): a configured provider that fails or produces no
# content must note a degradation, so the doc rides the degraded-ledger retry
# lane instead of being silently parked as "no text extracted".


class _FakeMedia:
    def __init__(self, transcript="", notes="", boom=False, exc=None):
        self._transcript = transcript
        self._notes = notes
        # `exc` lets a test pick the exception type (transient vs permanent);
        # `boom=True` keeps the legacy transient ConnectionError default.
        self._exc = exc or (ConnectionError("media host down") if boom else None)

    def transcribe_audio(self, file_path):
        if self._exc is not None:
            raise self._exc
        return self._transcript

    def analyze_video(self, file_path):
        if self._exc is not None:
            raise self._exc
        return self._notes


def test_extract_audio_provider_failure_notes_degradation():
    from extractors import begin_degradation_capture, collect_degradations, extract_audio

    begin_degradation_capture()
    extract_audio("/fake/voicemail.mp3", media_provider=_FakeMedia(boom=True))
    degs = collect_degradations()
    assert [d.reason for d in degs] == ["audio_extract_failed"]
    assert degs[0].transient is True  # ConnectionError is transient


def test_extract_audio_empty_transcript_notes_nothing():
    # The media provider is always wrapped: it returns "" ONLY for a
    # fallback-confirmed blank. extract_audio must treat that as clean.
    from extractors import begin_degradation_capture, collect_degradations, extract_audio

    begin_degradation_capture()
    extract_audio("/fake/voicemail.mp3", media_provider=_FakeMedia(transcript=""))
    assert collect_degradations() == []


def test_extract_audio_good_transcript_notes_nothing():
    from extractors import begin_degradation_capture, collect_degradations, extract_audio

    begin_degradation_capture()
    extract_audio("/fake/voicemail.mp3", media_provider=_FakeMedia(transcript="hello"))
    assert collect_degradations() == []


def test_extract_video_provider_failure_transient_notes_degradation():
    # A vision/video-provider OUTAGE is transient — a degradation (retries with
    # attempts unburned), NOT a permanent skip (was the bug).
    from extractors import (
        begin_degradation_capture,
        collect_degradations,
        collect_skips,
        extract_video,
    )

    begin_degradation_capture()
    extract_video("/fake/clip.mp4", media_provider=_FakeMedia(boom=True))
    degs = collect_degradations()
    assert [d.reason for d in degs] == ["video_extract_failed"]
    assert degs[0].transient is True  # ConnectionError is transient
    assert collect_skips() == []  # no longer a skip


def test_extract_video_provider_failure_permanent_notes_degradation():
    # An unclassified backend failure (bad response shape, codec the model
    # choked on) is non-transient — it rides the degraded ledger and caps
    # after the max attempts.
    from extractors import begin_degradation_capture, collect_degradations, extract_video

    begin_degradation_capture()
    extract_video(
        "/fake/clip.mp4", media_provider=_FakeMedia(exc=ValueError("bad shape"))
    )
    degs = collect_degradations()
    assert [d.reason for d in degs] == ["video_extract_failed"]
    assert degs[0].transient is False  # ValueError is not transient


def _oversize_provider(max_file_size_mb=0.001):
    """The real LiteLLM media provider with a tiny ceiling, so the production
    size guard — not a stand-in exception — decides the outcome."""
    from providers.media import _LiteLLMMediaProvider

    def _must_not_run(file_path):
        raise AssertionError("provider must not be called for an oversized file")

    return _LiteLLMMediaProvider(
        _must_not_run, _must_not_run, max_file_size_mb=max_file_size_mb
    )


def test_extract_video_oversized_is_skip_not_degradation(tmp_path):
    # #0481: a file over media.max_file_size_mb is a property of the FILE, not
    # of the backend — no retry can shrink it. Classifying it as a degradation
    # left it out of both LanceDB and the skip ledger, so the plain diff
    # re-admitted it every run and `attempts` ran to 113 against a cap of 5.
    from extractors import (
        begin_degradation_capture,
        collect_degradations,
        collect_skips,
        extract_video,
    )

    clip = tmp_path / "walkthrough.mp4"
    clip.write_bytes(b"\x00\x00\x00\x18ftypmp42" + b"\x00" * 4096)

    begin_degradation_capture()
    result = extract_video(clip, media_provider=_oversize_provider())
    assert result.full_text == ""
    assert collect_skips() == ["media_file_too_large"]
    assert collect_degradations() == []  # not a retry-forever degradation


def test_extract_audio_oversized_is_skip_not_degradation(tmp_path):
    from extractors import (
        begin_degradation_capture,
        collect_degradations,
        collect_skips,
        extract_audio,
    )

    voicemail = tmp_path / "voicemail.m4a"
    voicemail.write_bytes(b"\x00" * 4096)

    begin_degradation_capture()
    result = extract_audio(voicemail, media_provider=_oversize_provider())
    assert result.full_text == ""
    assert collect_skips() == ["media_file_too_large"]
    assert collect_degradations() == []


def test_extract_video_retrieval_stub_skips_without_provider_call(tmp_path):
    # A retrieval placeholder (a small JSON error envelope written in place of
    # the real video bytes, e.g. Zoho Cliq's attachment_access_time_expired) is
    # a permanent skip and must NOT reach the paid media provider — otherwise it
    # is re-downloaded/re-analyzed every indexing run forever (the #Maintenance
    # "no electric" video incident: video_extract_failed x9 and climbing).
    from extractors import (
        begin_degradation_capture,
        collect_degradations,
        collect_skips,
        extract_video,
    )

    stub = tmp_path / "clip.mp4"
    stub.write_text('{"code":"attachment_access_time_expired"}')

    class _Boom:
        def analyze_video(self, file_path):
            raise AssertionError("provider must not be called for a non-media stub")

    begin_degradation_capture()
    result = extract_video(stub, media_provider=_Boom())
    assert result.full_text == ""
    assert "media_retrieval_stub" in collect_skips()  # skip -> excluded from diff
    assert collect_degradations() == []  # not a retry-forever degradation


def test_extract_audio_retrieval_stub_skips_without_provider_call(tmp_path):
    from extractors import (
        begin_degradation_capture,
        collect_skips,
        extract_audio,
    )

    stub = tmp_path / "voicemail.m4a"
    stub.write_text('{"error":"link expired"}')

    class _Boom:
        def transcribe_audio(self, file_path):
            raise AssertionError("provider must not be called for a non-media stub")

    begin_degradation_capture()
    result = extract_audio(stub, media_provider=_Boom())
    assert result.full_text == ""
    assert "media_retrieval_stub" in collect_skips()


def test_extract_video_real_bytes_not_treated_as_stub(tmp_path):
    # Binary media that is not JSON must still reach the provider as normal.
    from extractors import begin_degradation_capture, collect_skips, extract_video

    real = tmp_path / "clip.mp4"
    real.write_bytes(b"\x00\x00\x00\x18ftypmp42" + b"\x00" * 128)  # mp4-ish binary

    class _OK:
        def analyze_video(self, file_path):
            return "a person waves at the camera"

    begin_degradation_capture()
    result = extract_video(real, media_provider=_OK())
    assert result.full_text == "a person waves at the camera"
    assert collect_skips() == []


def test_extract_video_empty_analysis_notes_nothing():
    # Confirmed-blank analysis (wrapper returns "") is clean — no note.
    from extractors import begin_degradation_capture, collect_degradations, extract_video

    begin_degradation_capture()
    extract_video("/fake/clip.mp4", media_provider=_FakeMedia(notes=""))
    assert collect_degradations() == []


# --- Dispatcher ---


def test_dispatch_markdown(tmp_path):
    """Dispatcher routes .md to extract_markdown."""
    from extractors import extract_text

    md_file = tmp_path / "test.md"
    md_file.write_text("# Dispatch test", encoding="utf-8")

    result = extract_text(md_file, ext="md")
    assert "Dispatch test" in result.full_text


def test_dispatch_pdf(tmp_path):
    """Dispatcher routes .pdf to extract_pdf."""
    from extractors import extract_text

    pdf_path = tmp_path / "test.pdf"
    _make_test_pdf(pdf_path, ["PDF dispatch content"])

    result = extract_text(pdf_path, ext="pdf")
    assert "PDF dispatch" in result.full_text


def test_dispatch_image():
    """Dispatcher routes .png to extract_image."""
    from extractors import extract_text
    from providers.ocr.base import OCRProvider

    class FakeOCR(OCRProvider):
        def extract(self, file_path, page=None):
            return "Image OCR text"

    result = extract_text("/fake/img.png", ext="png", ocr_provider=FakeOCR())
    assert "Image OCR text" in result.full_text


def test_dispatch_txt(tmp_path):
    """Dispatcher routes .txt to extract_plaintext."""
    from extractors import extract_text

    txt_file = tmp_path / "test.txt"
    txt_file.write_text("Plain text content", encoding="utf-8")

    result = extract_text(txt_file, ext="txt")
    assert "Plain text content" in result.full_text


def test_dispatch_xlsx(tmp_path):
    """Dispatcher routes .xlsx to extract_excel."""
    from extractors import extract_text
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Age"])
    ws.append(["Alice", 30])
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(str(xlsx_path))

    result = extract_text(xlsx_path, ext="xlsx")
    assert "Name" in result.full_text
    assert "Alice" in result.full_text


def test_dispatch_docx(tmp_path):
    """Dispatcher routes .docx to extract_markitdown."""
    from extractors import extract_text
    from unittest.mock import patch, MagicMock

    mock_result = MagicMock()
    mock_result.text_content = "# Converted Document\n\nBody text."
    mock_md = MagicMock()
    mock_md.convert.return_value = mock_result

    with patch("extractors._get_markitdown", return_value=mock_md):
        result = extract_text("/fake/doc.docx", ext="docx")

    assert "Converted Document" in result.full_text


def test_dispatch_csv(tmp_path):
    """Dispatcher routes .csv to extract_markitdown."""
    from extractors import extract_text
    from unittest.mock import patch, MagicMock

    mock_result = MagicMock()
    mock_result.text_content = "| col1 | col2 |\n|---|---|\n| a | b |"
    mock_md = MagicMock()
    mock_md.convert.return_value = mock_result

    with patch("extractors._get_markitdown", return_value=mock_md):
        result = extract_text("/fake/data.csv", ext="csv")

    assert "col1" in result.full_text


def test_dispatch_html():
    """Dispatcher routes .html to extract_markitdown."""
    from extractors import extract_text
    from unittest.mock import patch, MagicMock

    mock_result = MagicMock()
    mock_result.text_content = "# Page Title\n\nParagraph content."
    mock_md = MagicMock()
    mock_md.convert.return_value = mock_result

    with patch("extractors._get_markitdown", return_value=mock_md):
        result = extract_text("/fake/page.html", ext="html")

    assert "Page Title" in result.full_text


def test_dispatch_pptx():
    """Dispatcher routes .pptx to extract_markitdown."""
    from extractors import extract_text
    from unittest.mock import patch, MagicMock

    mock_result = MagicMock()
    mock_result.text_content = "# Slide 1\n\nBullet point."
    mock_md = MagicMock()
    mock_md.convert.return_value = mock_result

    with patch("extractors._get_markitdown", return_value=mock_md):
        result = extract_text("/fake/deck.pptx", ext="pptx")

    assert "Slide 1" in result.full_text


def test_dispatch_unknown_ext(tmp_path):
    """Unknown extension returns empty text."""
    from extractors import extract_text

    result = extract_text("/fake/file.xyz", ext="xyz")
    assert result.full_text == ""


# --- Plaintext extraction ---


def test_extract_plaintext_basic(tmp_path):
    """Plaintext extraction reads file content."""
    from extractors import extract_plaintext

    txt_file = tmp_path / "test.txt"
    txt_file.write_text("Hello, world!", encoding="utf-8")

    result = extract_plaintext(txt_file)
    assert result.full_text == "Hello, world!"
    assert result.frontmatter == {}


def test_extract_plaintext_unicode(tmp_path):
    """Plaintext handles Unicode characters."""
    from extractors import extract_plaintext

    txt_file = tmp_path / "unicode.txt"
    txt_file.write_text("日本語テスト café 🚀", encoding="utf-8")

    result = extract_plaintext(txt_file)
    assert "日本語テスト" in result.full_text
    assert "café" in result.full_text


def test_extract_plaintext_empty(tmp_path):
    """Empty plaintext file returns empty text."""
    from extractors import extract_plaintext

    txt_file = tmp_path / "empty.txt"
    txt_file.write_text("", encoding="utf-8")

    result = extract_plaintext(txt_file)
    assert result.full_text == ""


# --- Excel extraction ---


def test_extract_excel_text_only(tmp_path):
    """Excel extraction includes text cells and skips numbers."""
    from extractors import extract_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Score", "Notes"])
    ws.append(["Alice", 95, "Excellent"])
    ws.append(["Bob", 80, "Good"])
    ws.append([None, 100, None])
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(str(xlsx_path))

    result = extract_excel(xlsx_path)
    assert "Sheet: Data" in result.full_text
    assert "Headers: Name | Score | Notes" in result.full_text
    assert "Alice" in result.full_text
    assert "Excellent" in result.full_text
    assert "Bob" in result.full_text
    assert "Good" in result.full_text
    # Numbers should NOT appear as standalone extracted values (only in headers)
    lines = result.full_text.split("\n")
    data_lines = [l for l in lines if not l.startswith(("Sheet:", "Headers:"))]
    for line in data_lines:
        assert "95" not in line
        assert "80" not in line
        assert "100" not in line


def test_extract_excel_headers_always_included(tmp_path):
    """Excel headers are always included even if they contain numbers."""
    from extractors import extract_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([2024, "Q1", "Revenue"])
    ws.append([None, "Good quarter", None])
    xlsx_path = tmp_path / "headers.xlsx"
    wb.save(str(xlsx_path))

    result = extract_excel(xlsx_path)
    assert "Headers: 2024 | Q1 | Revenue" in result.full_text


def test_extract_excel_multiple_sheets(tmp_path):
    """Excel extraction handles multiple sheets."""
    from extractors import extract_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Header1"])
    ws1.append(["Value1"])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Header2"])
    ws2.append(["Value2"])

    xlsx_path = tmp_path / "multi.xlsx"
    wb.save(str(xlsx_path))

    result = extract_excel(xlsx_path)
    assert "Sheet: Sheet1" in result.full_text
    assert "Sheet: Sheet2" in result.full_text
    assert "Value1" in result.full_text
    assert "Value2" in result.full_text


def test_extract_excel_empty_workbook(tmp_path):
    """Empty Excel workbook returns empty text."""
    from extractors import extract_excel
    import openpyxl

    wb = openpyxl.Workbook()
    # Default sheet exists but has no rows
    xlsx_path = tmp_path / "empty.xlsx"
    wb.save(str(xlsx_path))

    result = extract_excel(xlsx_path)
    # Empty sheet has no rows via iter_rows in read_only mode
    assert result.frontmatter == {}


# --- MarkItDown extraction ---


def test_extract_markitdown_mock():
    """MarkItDown conversion works with a mock."""
    from extractors import extract_markitdown
    from unittest.mock import patch, MagicMock

    mock_result = MagicMock()
    mock_result.text_content = "# Document Title\n\nConverted content here."
    mock_md = MagicMock()
    mock_md.convert.return_value = mock_result

    with patch("extractors._get_markitdown", return_value=mock_md):
        result = extract_markitdown("/fake/doc.docx")

    assert "Document Title" in result.full_text
    assert "Converted content" in result.full_text


def test_extract_markitdown_not_installed():
    """Graceful fallback when markitdown is not installed."""
    from extractors import extract_markitdown
    from unittest.mock import patch

    with patch("extractors._get_markitdown", return_value=None):
        result = extract_markitdown("/fake/doc.docx")

    assert result.full_text == ""


def test_extract_markitdown_conversion_error():
    """Graceful fallback when conversion fails."""
    from extractors import extract_markitdown
    from unittest.mock import patch, MagicMock

    mock_md = MagicMock()
    mock_md.convert.side_effect = RuntimeError("Conversion failed")

    with patch("extractors._get_markitdown", return_value=mock_md):
        result = extract_markitdown("/fake/doc.docx")

    assert result.full_text == ""


def test_fallback_chain_recovers_content_through_extract_image():
    """Primary describe reachable-but-empty + a fallback that returns text ->
    extract_image indexes the RECOVERED text and notes NO degradation (clean)."""
    from extractors import begin_degradation_capture, collect_degradations, extract_image
    from providers.ocr.base import OCRProvider
    from providers.ocr.fallback import FallbackOCRProvider

    class EmptyPrimary(OCRProvider):
        def extract(self, file_path, page=None): return ""
        def describe(self, file_path): return ""      # reachable but empty

    wrapped = FallbackOCRProvider(EmptyPrimary(),
                                  describe_fallback=lambda p: "a recovered description")
    begin_degradation_capture()
    result = extract_image("/fake/path.png", ocr_provider=wrapped)
    assert "a recovered description" in result.full_text   # recovered content indexed
    assert collect_degradations() == []                    # clean, no degradation


def test_fallback_chain_confirmed_blank_is_clean_through_extract_image():
    """Primary empty + fallback ALSO empty -> confirmed blank -> extract_image
    indexes only the metadata stub and notes NO degradation (doc drops from ledger)."""
    from extractors import begin_degradation_capture, collect_degradations, extract_image
    from providers.ocr.base import OCRProvider
    from providers.ocr.fallback import FallbackOCRProvider

    class EmptyPrimary(OCRProvider):
        def extract(self, file_path, page=None): return ""
        def describe(self, file_path): return ""

    wrapped = FallbackOCRProvider(EmptyPrimary(),
                                  describe_fallback=lambda p: "")   # fallback agrees: blank
    begin_degradation_capture()
    extract_image("/fake/path.png", ocr_provider=wrapped)
    assert collect_degradations() == []                    # confirmed blank == clean


def test_fallback_chain_dark_mode_outage_degrades_transient_through_extract_image():
    """No fallback configured + primary empty -> wrapper raises transient ->
    extract_image notes ocr_describe_failed as TRANSIENT (retries, never capped)."""
    from extractors import begin_degradation_capture, collect_degradations, extract_image
    from providers.ocr.base import OCRProvider
    from providers.ocr.fallback import FallbackOCRProvider

    class EmptyPrimary(OCRProvider):
        def extract(self, file_path, page=None): return ""
        def describe(self, file_path): return ""

    wrapped = FallbackOCRProvider(EmptyPrimary(), describe_fallback=None)  # dark
    begin_degradation_capture()
    extract_image("/fake/path.png", ocr_provider=wrapped)
    degs = collect_degradations()
    assert [d.reason for d in degs] == ["ocr_describe_failed"]
    assert degs[0].transient is True                       # transient -> #60 never caps
