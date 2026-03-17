"""Text extraction from Markdown, PDF, and image files.

PDF extraction uses PyMuPDF (fitz) for native text, with optional OCR fallback.
PDF metadata (title, author, dates) is extracted via PyMuPDF and stored in frontmatter.

Image extraction uses OCR provider's describe() for rich text + visual description.
Image metadata (EXIF: camera, date, GPS, dimensions) is extracted via Pillow.

Markdown extraction parses YAML frontmatter (tags, status, title, created, etc.)
and strips it from the indexed content so embeddings aren't polluted by YAML syntax.
"""

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from providers.ocr.base import OCRProvider

logger = logging.getLogger(__name__)


@dataclass
class PageText:
    """Text extracted from a single page (or whole file for Markdown/images)."""
    page: int          # 0-based page number (0 for MD/images)
    text: str
    was_ocr: bool = False


@dataclass
class ExtractionResult:
    """Result of extracting text from a document."""
    pages: list[PageText] = field(default_factory=list)
    full_text: str = ""
    frontmatter: dict = field(default_factory=dict)

    @staticmethod
    def from_pages(
        pages: list[PageText], frontmatter: dict | None = None,
    ) -> "ExtractionResult":
        full = "\n\n".join(p.text for p in pages if p.text.strip())
        return ExtractionResult(pages=pages, full_text=full, frontmatter=frontmatter or {})

    @staticmethod
    def from_text(text: str, frontmatter: dict | None = None) -> "ExtractionResult":
        return ExtractionResult(
            pages=[PageText(page=0, text=text)],
            full_text=text,
            frontmatter=frontmatter or {},
        )


# ---------------------------------------------------------------------------
# Frontmatter + title helpers
# ---------------------------------------------------------------------------

def parse_frontmatter(content: str) -> tuple[dict, str]:
    """Parse YAML frontmatter from markdown content.

    Returns (frontmatter_dict, content_without_frontmatter).
    If no valid frontmatter is found, returns ({}, original_content).

    Handles Obsidian-style frontmatter:
        ---
        tags: [recipe, korean]
        status: active
        created: 2026-01-15
        ---
    """
    if not content.startswith("---"):
        return {}, content

    match = re.match(r'^---\s*\n(.*?)\n---\s*\n?', content, re.DOTALL)
    if not match:
        return {}, content

    yaml_block = match.group(1)
    body = content[match.end():]

    try:
        import yaml
        fm = yaml.safe_load(yaml_block)
        if not isinstance(fm, dict):
            return {}, content
    except Exception:
        return {}, content

    return fm, body


def extract_title(content: str, doc_id: str) -> str:
    """Extract a document title.

    Priority:
      1. First # heading in the content
      2. Filename without extension (from doc_id)
    """
    match = re.search(r'^#\s+(.+)$', content, re.MULTILINE)
    if match:
        return match.group(1).strip()
    return Path(doc_id).stem


def derive_folder(doc_id: str) -> str:
    """Derive the top-level folder from a vault-relative doc_id.

    Examples:
      "subfolder/recipe.md"  → "subfolder"
      "Archive/old_note.md"  → "Archive"
      "note1.md"             → ""  (root)
    """
    parts = doc_id.split("/")
    return parts[0] if len(parts) > 1 else ""


def normalize_tags(raw_tags) -> str:
    """Normalize frontmatter tags to a comma-separated string.

    Handles:
      - list: [recipe, korean]  → "recipe,korean"
      - string: "recipe, korean" → "recipe,korean"
      - None → ""
    """
    if raw_tags is None:
        return ""
    if isinstance(raw_tags, list):
        return ",".join(str(t).strip() for t in raw_tags if t)
    if isinstance(raw_tags, str):
        return ",".join(t.strip() for t in raw_tags.split(",") if t.strip())
    return str(raw_tags)


# ---------------------------------------------------------------------------
# Markdown
# ---------------------------------------------------------------------------

def extract_markdown(file_path: str | Path) -> ExtractionResult:
    """Read a Markdown file, parse frontmatter, and return text without YAML block.

    Frontmatter (tags, status, created, etc.) is stored in result.frontmatter.
    The YAML block is stripped from full_text so embeddings aren't polluted.
    """
    with open(file_path, encoding="utf-8", errors="replace") as f:
        raw = f.read()
    frontmatter, body = parse_frontmatter(raw)
    return ExtractionResult.from_text(body, frontmatter=frontmatter)


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def _extract_pdf_metadata(doc) -> dict:
    """Extract document-level metadata from a PyMuPDF document.

    Returns a dict suitable for storing in ExtractionResult.frontmatter.
    """
    meta: dict = {}
    try:
        info = doc.metadata or {}
        if info.get("title"):
            meta["title"] = info["title"].strip()
        if info.get("author"):
            meta["author"] = info["author"].strip()
        if info.get("subject"):
            meta["subject"] = info["subject"].strip()
        if info.get("creator"):
            meta["pdf_creator"] = info["creator"].strip()
        if info.get("producer"):
            meta["pdf_producer"] = info["producer"].strip()
        if info.get("creationDate"):
            meta["created"] = _parse_pdf_date(info["creationDate"])
        if info.get("modDate"):
            meta["modified"] = _parse_pdf_date(info["modDate"])
        meta["page_count"] = len(doc)
    except Exception as e:
        logger.debug("Could not extract PDF metadata: %s", e)
    return meta


def _parse_pdf_date(raw: str) -> str:
    """Parse PDF date format (D:20240115120000+00'00') into readable string."""
    if not raw:
        return ""
    raw = raw.strip()
    if raw.startswith("D:"):
        raw = raw[2:]
    # Take the first 14 chars: YYYYMMDDHHMMSS
    digits = re.sub(r"[^\d]", "", raw[:14])
    if len(digits) >= 8:
        y, m, d = digits[:4], digits[4:6], digits[6:8]
        result = f"{y}-{m}-{d}"
        if len(digits) >= 14:
            result += f" {digits[8:10]}:{digits[10:12]}:{digits[12:14]}"
        return result
    return raw


def _format_pdf_metadata_header(meta: dict) -> str:
    """Format PDF metadata as a text header to prepend to the first page."""
    lines = []
    if meta.get("title"):
        lines.append(f"Document title: {meta['title']}")
    if meta.get("author"):
        lines.append(f"Author: {meta['author']}")
    if meta.get("subject"):
        lines.append(f"Subject: {meta['subject']}")
    if meta.get("created"):
        lines.append(f"Created: {meta['created']}")
    if meta.get("modified"):
        lines.append(f"Modified: {meta['modified']}")
    if meta.get("pdf_creator"):
        lines.append(f"Creator software: {meta['pdf_creator']}")
    if meta.get("page_count"):
        lines.append(f"Pages: {meta['page_count']}")
    return "\n".join(lines)


def extract_pdf(
    file_path: str | Path,
    strategy: str = "text_then_ocr",
    ocr_provider: Optional[OCRProvider] = None,
    min_text_chars: int = 200,
    ocr_page_limit: int = 200,
) -> ExtractionResult:
    """Extract text and metadata from a PDF using PyMuPDF.

    Strategies:
        text_only      — extract embedded text only
        text_then_ocr  — extract text; if a page has < min_text_chars, OCR it
        ocr_only       — OCR every page (ignore embedded text)

    PDF metadata (title, author, creation date, page count) is extracted and
    stored in result.frontmatter. A metadata header is prepended to the first page.
    """
    import fitz  # PyMuPDF

    doc = fitz.open(str(file_path))
    pdf_meta = _extract_pdf_metadata(doc)
    pages: list[PageText] = []

    for page_num in range(len(doc)):
        page = doc[page_num]

        if strategy == "ocr_only":
            text = _ocr_page(doc, page_num, ocr_provider, ocr_page_limit, len(pages))
            pages.append(PageText(page=page_num, text=text, was_ocr=True))

        elif strategy == "text_then_ocr":
            text = page.get_text("text").strip()
            was_ocr = False
            if len(text) < min_text_chars and ocr_provider is not None:
                ocr_text = _ocr_page(doc, page_num, ocr_provider, ocr_page_limit, _ocr_count(pages))
                if ocr_text.strip():
                    text = ocr_text
                    was_ocr = True
            pages.append(PageText(page=page_num, text=text, was_ocr=was_ocr))

        else:  # text_only
            text = page.get_text("text").strip()
            pages.append(PageText(page=page_num, text=text))

    doc.close()

    # Prepend metadata header to first page content so it gets embedded/searchable
    meta_header = _format_pdf_metadata_header(pdf_meta)
    if meta_header and pages:
        pages[0] = PageText(
            page=pages[0].page,
            text=meta_header + "\n\n" + pages[0].text,
            was_ocr=pages[0].was_ocr,
        )

    return ExtractionResult.from_pages(pages, frontmatter=pdf_meta)


def _ocr_count(pages: list[PageText]) -> int:
    """Count how many pages have been OCR'd so far."""
    return sum(1 for p in pages if p.was_ocr)


def _ocr_page(
    doc,  # fitz.Document
    page_num: int,
    ocr_provider: Optional[OCRProvider],
    ocr_page_limit: int,
    ocr_so_far: int,
) -> str:
    """Render a PDF page to a PNG image and OCR it."""
    if ocr_provider is None:
        return ""
    if ocr_so_far >= ocr_page_limit:
        return ""

    import tempfile

    page = doc[page_num]
    # Render page at 2x resolution for better OCR quality
    pix = page.get_pixmap(dpi=200)
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        pix.save(tmp.name)
        tmp_path = tmp.name

    try:
        text = ocr_provider.extract(tmp_path)
    except Exception as e:
        logger.warning("OCR failed for page %d: %s", page_num, e)
        text = ""
    finally:
        Path(tmp_path).unlink(missing_ok=True)

    return text


# ---------------------------------------------------------------------------
# Image metadata (EXIF)
# ---------------------------------------------------------------------------

def _extract_image_metadata(file_path: str | Path) -> dict:
    """Extract metadata from an image file using Pillow.

    Returns a dict with available fields: dimensions, camera, date_taken,
    gps_lat, gps_lon, software, orientation.
    """
    meta: dict = {}
    try:
        from PIL import Image
        from PIL.ExifTags import TAGS, GPSTAGS

        img = Image.open(file_path)
        meta["width"] = img.width
        meta["height"] = img.height
        meta["format"] = img.format or ""

        exif_data = img.getexif()
        if not exif_data:
            return meta

        decoded: dict = {}
        for tag_id, value in exif_data.items():
            tag_name = TAGS.get(tag_id, str(tag_id))
            decoded[tag_name] = value

        if "Make" in decoded and "Model" in decoded:
            meta["camera"] = f"{decoded['Make']} {decoded['Model']}".strip()
        elif "Model" in decoded:
            meta["camera"] = str(decoded["Model"]).strip()

        if "DateTime" in decoded:
            meta["date_taken"] = str(decoded["DateTime"])
        elif "DateTimeOriginal" in decoded:
            meta["date_taken"] = str(decoded["DateTimeOriginal"])

        if "Software" in decoded:
            meta["software"] = str(decoded["Software"])

        # GPS coordinates
        gps_ifd = exif_data.get_ifd(0x8825)  # GPSInfo IFD
        if gps_ifd:
            gps_decoded = {}
            for tag_id, value in gps_ifd.items():
                tag_name = GPSTAGS.get(tag_id, str(tag_id))
                gps_decoded[tag_name] = value

            def _dms_to_decimal(dms, ref):
                d, m, s = float(dms[0]), float(dms[1]), float(dms[2])
                decimal = d + m / 60.0 + s / 3600.0
                if ref in ("S", "W"):
                    decimal = -decimal
                return round(decimal, 6)

            if "GPSLatitude" in gps_decoded and "GPSLatitudeRef" in gps_decoded:
                meta["gps_lat"] = _dms_to_decimal(
                    gps_decoded["GPSLatitude"], gps_decoded["GPSLatitudeRef"]
                )
            if "GPSLongitude" in gps_decoded and "GPSLongitudeRef" in gps_decoded:
                meta["gps_lon"] = _dms_to_decimal(
                    gps_decoded["GPSLongitude"], gps_decoded["GPSLongitudeRef"]
                )

    except Exception as e:
        logger.warning("Could not extract image metadata from %s: %s", file_path, e)

    return meta


def _format_image_metadata_header(meta: dict) -> str:
    """Format image metadata as a text header to prepend to extracted content."""
    lines = []
    if meta.get("width") and meta.get("height"):
        lines.append(f"Image dimensions: {meta['width']}x{meta['height']}")
    if meta.get("format"):
        lines.append(f"Format: {meta['format']}")
    if meta.get("camera"):
        lines.append(f"Camera: {meta['camera']}")
    if meta.get("date_taken"):
        lines.append(f"Date taken: {meta['date_taken']}")
    if meta.get("software"):
        lines.append(f"Software: {meta['software']}")
    if meta.get("gps_lat") is not None and meta.get("gps_lon") is not None:
        lines.append(f"GPS: {meta['gps_lat']}, {meta['gps_lon']}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Image
# ---------------------------------------------------------------------------

def extract_image(
    file_path: str | Path,
    ocr_provider: Optional[OCRProvider] = None,
) -> ExtractionResult:
    """Extract text, visual description, and metadata from an image file.

    Uses OCR provider's describe() for rich content (text + visual description).
    Extracts EXIF metadata (camera, date, GPS, dimensions) and stores in frontmatter.
    """
    file_path = Path(file_path)
    meta = _extract_image_metadata(file_path)
    fm: dict = {}
    if meta.get("date_taken"):
        fm["created"] = meta["date_taken"]
    if meta.get("camera"):
        fm["camera"] = meta["camera"]
    if meta.get("gps_lat") is not None:
        fm["gps_lat"] = meta["gps_lat"]
    if meta.get("gps_lon") is not None:
        fm["gps_lon"] = meta["gps_lon"]
    if meta.get("width") and meta.get("height"):
        fm["dimensions"] = f"{meta['width']}x{meta['height']}"

    if ocr_provider is None:
        header = _format_image_metadata_header(meta)
        return ExtractionResult.from_text(header, frontmatter=fm)

    try:
        vision_text = ocr_provider.describe(str(file_path))
    except Exception as e:
        logger.warning("OCR describe failed for %s: %s", file_path, e)
        vision_text = ""
    header = _format_image_metadata_header(meta)
    parts = [p for p in [header, vision_text] if p.strip()]
    full_text = "\n\n".join(parts)
    return ExtractionResult.from_text(full_text, frontmatter=fm)


# ---------------------------------------------------------------------------
# MarkItDown (docx, doc, pptx, html, epub, rtf, csv)
# ---------------------------------------------------------------------------

_markitdown_instance = None


def _get_markitdown():
    """Lazy-init singleton. Returns None if markitdown not installed."""
    global _markitdown_instance
    if _markitdown_instance is not None:
        return _markitdown_instance
    try:
        from markitdown import MarkItDown
        _markitdown_instance = MarkItDown()
        return _markitdown_instance
    except ImportError:
        logger.warning("markitdown not installed — install with: pip install 'markitdown[all]'")
        return None


def extract_markitdown(file_path: str | Path) -> ExtractionResult:
    """Convert a document to Markdown via MarkItDown and return as ExtractionResult.

    Supports: docx, doc, pptx, html, htm, epub, rtf, csv.
    """
    md = _get_markitdown()
    if md is None:
        return ExtractionResult.from_text("")

    try:
        result = md.convert(str(file_path))
        text = result.text_content or ""
    except Exception as e:
        logger.warning("MarkItDown conversion failed for %s: %s", file_path, e)
        return ExtractionResult.from_text("")

    return ExtractionResult.from_text(text)


# ---------------------------------------------------------------------------
# Excel (xlsx, xls) — text-only extraction
# ---------------------------------------------------------------------------

def extract_excel(file_path: str | Path) -> ExtractionResult:
    """Extract text-only cells from an Excel workbook.

    Headers (first row) are always included. Remaining rows include only
    cells whose value is a string — numbers, dates, and empty cells are skipped.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed — install with: pip install openpyxl")
        return ExtractionResult.from_text("")

    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Failed to open workbook %s: %s", file_path, e)
        return ExtractionResult.from_text("")

    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows())
            if not rows:
                continue

            section_lines: list[str] = [f"Sheet: {sheet_name}"]

            # First row → headers (always included regardless of type)
            header_cells = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
            section_lines.append("Headers: " + " | ".join(header_cells))

            # Remaining rows → text-only cells
            for row in rows[1:]:
                text_cells = [
                    str(cell.value).strip()
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.strip()
                ]
                if text_cells:
                    section_lines.append(" | ".join(text_cells))

            parts.append("\n".join(section_lines))
    finally:
        wb.close()

    full_text = "\n\n".join(parts)
    return ExtractionResult.from_text(full_text)


# ---------------------------------------------------------------------------
# Plain text (txt)
# ---------------------------------------------------------------------------

def extract_plaintext(file_path: str | Path) -> ExtractionResult:
    """Read a plain text file (UTF-8, with replacement for bad bytes)."""
    with open(file_path, encoding="utf-8", errors="replace") as f:
        text = f.read()
    return ExtractionResult.from_text(text)


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

def extract_text(
    file_path: str | Path,
    ext: str,
    ocr_provider: Optional[OCRProvider] = None,
    pdf_strategy: str = "text_then_ocr",
    min_text_chars: int = 200,
    ocr_page_limit: int = 200,
) -> ExtractionResult:
    """Route to the right extractor based on file extension."""
    if ext == "md":
        return extract_markdown(file_path)
    elif ext == "pdf":
        return extract_pdf(
            file_path,
            strategy=pdf_strategy,
            ocr_provider=ocr_provider,
            min_text_chars=min_text_chars,
            ocr_page_limit=ocr_page_limit,
        )
    elif ext in ("png", "jpg", "jpeg", "gif", "webp"):
        return extract_image(file_path, ocr_provider=ocr_provider)
    elif ext == "txt":
        return extract_plaintext(file_path)
    elif ext in ("xlsx", "xls"):
        return extract_excel(file_path)
    elif ext in ("docx", "doc", "pptx", "html", "htm", "epub", "rtf", "csv"):
        return extract_markitdown(file_path)
    else:
        return ExtractionResult.from_text("")
